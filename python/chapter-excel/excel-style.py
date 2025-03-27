from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Side, Border, PatternFill, Alignment, GradientFill

# Load the workbook
# wb = load_workbook(filename='example.xlsx')
# sheet = wb['Sheet2']


# part 1: Set the font style
# cell = sheet['A1']
# print(f'cell A1 font: {cell.font}')

# font = Font(name='Arial', size=12, bold=True, italic=True, color=Color(rgb='FF0000'))
# cell.font = font
# print(f'cell A1 font after setting: {cell.font}')

# wb.save(filename='example.xlsx')



# part 2: set the border style
# side = Side(style='thick', color='FF0000')
# border = Border(left=side, right=side, top=side, bottom=side)

# cells = sheet[1]
# print(f'cell 0 border: {cells[0].border}')

# for cell in cells:
#   cell.border = border

# wb.save(filename='example.xlsx')



# part 3: set the background color
# pattern_fill = PatternFill(patternType='solid', fgColor='FF0000')
# gradient_fill = GradientFill(stop=("000000", "FFFFFF"))

# cells = sheet[3]
# for cell in cells:
#   cell.fill = pattern_fill

# cells = sheet[4]
# for cell in cells:
#   cell.fill = gradient_fill

# wb.save(filename='example.xlsx')



# part 4: set the alignment style
# alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
# cells = sheet[5]
# for cell in cells:
#   cell.alignment = alignment

# wb.save(filename='example.xlsx')



# part 5: set line height and column width
# sheet.row_dimensions[1].height = 30
# sheet.column_dimensions['A'].width = 20
# wb.save(filename='example.xlsx')



# part 6: merge cells
# 合并后的单元格只会显示合并区域中最右上角的单元格的值，会导致其他单元格内容丢失
# sheet.merge_cells('C9:C10')
# sheet.unmerge_cells('C9:C10')
# wb.save(filename='example.xlsx')


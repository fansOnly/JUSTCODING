import openpyxl

# load workbook
wb = openpyxl.load_workbook('example.xlsx')


# get sheet names
sheetnames = wb.sheetnames
print(sheetnames)

# part 1: get active sheet
# active_sheet = wb.active
# print(f'active sheet: {active_sheet}')
# print(f'active sheet name: {active_sheet.title}')


# part 2: get sheet by name
# sheet = wb['黔秀普卡号']
# print(f'sheet: {sheet}')
# print(f'sheet type: {type(sheet)}')
# print(f'sheet name: {sheet.title}')
# print(f'sheet rows: {sheet.max_row}')
# print(f'sheet columns: {sheet.max_column}')
# print(f'sheet.dimensions: {sheet.dimensions}')
# print(f'sheet.max_row: {sheet.max_row}')
# print(f'sheet.max_column: {sheet.max_column}')

# part 3: get cell value
# sheet = wb['黔秀普卡号']
# cell_b1 = sheet['B1']
# print(f'cell B1: {cell_b1}')
# print(f'cell B1 value: {cell_b1.value}')
# print(f'cell B1 Row: {cell_b1.row}, Column: {cell_b1.column}')
# # print cell coordinate and value
# print(f'Cell {cell_b1.coordinate} is {cell_b1.value}')

# part 4: print odd rows in column B
# for i in range(1, 10, 2):
#   print(i, sheet.cell(row=i, column=2).value)


# part 5: get multiple cells = ['A1:C5']
# cells = sheet['A1:C5']
# print(f'cells type: {type(cells)} \n')

# for rows in cells:
#   for cell in rows:
#     print(cell.value, end='|')
#   print('\n')

# # use inter_rows
# rows = sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3)
# for row in rows:
#   for cell in row:
#     print(cell.value, end='|')
#   print('\n')

# use inter_cols
# cols = sheet.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3)
# for col in cols:
#   for cell in col:
#     print(cell.value, end='|')
#   print('\n')


# practice 1
# 找出 example.xlsx 中 Sheet2 表中空着的格子，并输出这些格子的坐标
# sheet2 = wb['Sheet2']
# print(f'Sheet2 dimensions: {sheet2.dimensions}')
# cells = sheet2[sheet2.dimensions]

# for rows in cells:
#   for cell in rows:
#     if not cell.value:
#       print(f'Cell {cell.coordinate} is empty \n')


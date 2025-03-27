from openpyxl import load_workbook
from openpyxl import Workbook
import warnings
warnings.filterwarnings("ignore")


# part 1: edit value in original file
# wb = load_workbook(filename='example.xlsx')
# sheet = wb['Sheet4']
# print(f'before update: {sheet["A1"].value}')
# sheet["A1"].value = "yy"
# print(f'after update: {sheet["A1"].value}')
# wb.save(filename='example.xlsx')

# # save new file
# # wb.save(filename='example-1.xlsx')



# part 2: write data to a new file
# wb = Workbook()
# print(f'default sheet name: {wb.sheetnames}')
# sheet = wb.create_sheet(title='Sheet1', index=0)
# print(f'current sheet name: {wb.sheetnames}')
# sheet['A1'] = 25
# sheet['B1'] = 67

# # write formula to a cell
# sheet['C1'] = '=SUM(A1:B1)'
# wb.save(filename='example2.xlsx')

# # validate formula
# wb2 = load_workbook(filename='example2.xlsx', data_only=True)
# sheet2 = wb2['Sheet1']
# print(f'formula result: {sheet2["C1"].value}')



# part 3: insert column and row
# wb = load_workbook(filename='example.xlsx')
# sheet = wb['Sheet5']

# insert column
# print(f'before insert column: {sheet.max_column}')
# sheet.insert_cols(idx=2, amount=1)
# print(f'after insert column: {sheet.max_column}')
# # wb.save(filename='example.xlsx')

# insert row
# print(f'before insert row: {sheet.max_row}')
# sheet.insert_rows(idx=2, amount=3)
# print(f'after insert row: {sheet.max_row}')
# # wb.save(filename='example.xlsx')

# delete column
# sheet.delete_cols(idx=2, amount=1)
# sheet.delete_rows(idx=2, amount=3)
# wb.save(filename='example.xlsx')

# move cell
# sheet.move_range('A1:B1', rows=2, cols=5)
# wb.save(filename='example.xlsx')
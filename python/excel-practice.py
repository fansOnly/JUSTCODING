from openpyxl import load_workbook

# practice 1
# 将 example.xlsx 中的 Sheet1 拆分成两个新的工作表，客户信息表(A-E)，卡信息表(F-H)
# wb = load_workbook(filename='practice/example.xlsx')
# sheet = wb['黔秀普卡号']
# customer_sheet = wb.create_sheet(title='客户信息表')
# card_sheet = wb.create_sheet(title='卡信息表')

# # 复制客户信息表
# for row in range(1, sheet.max_row+1):
#   for col in range(1, 6):
#     customer_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value

# print(f'{customer_sheet.title} 工作表已生成: {customer_sheet.dimensions}')

# # 复制卡信息表
# for row in range(1, sheet.max_row+1):
#   for col in range(6, sheet.max_column+1):
#     card_sheet.cell(row=row, column=col-5).value = sheet.cell(row=row, column=col).value

# print(f'{card_sheet.title} 工作表已生成: {card_sheet.dimensions}')

# # 保存文件
# wb.save(filename='practice/example.xlsx')



# practice 2
# 将 example.xlsx 中的客户信息表(A-E)，卡信息表(F-H)合并成一张新表，客户卡信息表(A-H)
# wb = load_workbook(filename='practice/example.xlsx')
# customer_sheet = wb['客户信息表']
# card_sheet = wb['卡信息表']

# # 合并工作表
# merged_sheet = wb.create_sheet(title='客户卡信息表')
# for row in range(1, customer_sheet.max_row + 1):
#   for col in range(1, customer_sheet.max_column + 1):
#     merged_sheet.cell(row=row, column=col).value = customer_sheet.cell(row=row, column=col).value

# for row in range(1, card_sheet.max_row + 1):
#   for col in range(1, card_sheet.max_column + 1):
#     merged_sheet.cell(row=row, column=col+customer_sheet.max_column).value = card_sheet.cell(row=row, column=col).value

# print(f'{merged_sheet.title} 工作表已生成: {merged_sheet.dimensions}')

# # 保存文件
# wb.save(filename='practice/example.xlsx')